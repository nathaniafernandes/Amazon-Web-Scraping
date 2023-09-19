import requests
from bs4 import BeautifulSoup
import openpyxl  
import time
import random

# User-Agent strings to mimic different browsers
user_agents = [
    # Popular Desktop Browsers
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Firefox/86.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/88.0.705.63',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Firefox/50.0',

    # Popular Mobile Browsers
    'Mozilla/5.0 (Linux; Android 10; SM-A505F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.101 Mobile Safari/537.36',
    'Mozilla/5.0 (iPhone; CPU iPhone OS 14_4 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1',
    'Mozilla/5.0 (iPad; CPU OS 14_4 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1',

    # Popular Web Crawlers
    'Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)',
    'Mozilla/5.0 (compatible; Bingbot/2.0; +http://www.bing.com/bingbot.htm)',

    # Common Libraries
    'Mozilla/5.0 (compatible; Konqueror/3.5; Linux) KHTML/3.5.5 (like Gecko) (Kubuntu)',
    'Mozilla/5.0 (X11; U; Linux x86_64; en-US; rv:1.9.0.3) Gecko/2008092416 Firefox/3.0.3',

    # Additional User Agents
    'Opera/9.80 (Windows NT 6.2; WOW64) Presto/2.12.388 Version/12.17',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; Media Center PC',
    'Mozilla/4.08 (compatible; MSIE 6.0; Windows NT 5.1)',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
]

# Function to send a request with random User-Agent and delay
def send_request(url):
    headers = {
        'User-Agent': random.choice(user_agents),
    }
    response = requests.get(url, headers=headers)
    time.sleep(random.uniform(0.5, 2))  # Add a random delay between requests
    return response

url = input("Input your URL: ")


# Try to send the request multiple times with different user agents
for _ in range(3):
    response = send_request(url)
    if response.status_code == 200:
        break
else:
    print("Failed to fetch the page. Check your code and proxies if used.")
    exit(1)

soup = BeautifulSoup(response.text, 'html.parser')

# Find product title, price, and data-asin attributes
product_divs = soup.find_all('div', {'data-component-type': 's-search-result'})

data_asins = []
product_titles = []
prices = []


for product_div in product_divs:
    data_asin = product_div.get('data-asin', 'N/A')  
    product_title_elem = product_div.find('h2').find('a')

    if product_title_elem:
        product_title = product_title_elem.get_text(strip=True)
    else:
        # Check if the specified span tag is available when product_title_elem is not available
        product_title_span = product_div.find('span', {'class': 'a-size-base-plus a-color-base a-text-normal'})
        product_title = product_title_span.get_text(strip=True) if product_title_span else 'N/A'  # Use 'N/A' if neither title nor span is found

    price_elem = product_div.find('span', {'class': 'a-price-whole'})
    price = price_elem.get_text(strip=True) if price_elem else 'N/A'  # Use 'N/A' if price is not found

    data_asins.append(data_asin)
    product_titles.append(product_title)
    prices.append(price)

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet.cell(row=1, column=1, value="Data-ASIN")
worksheet.cell(row=1, column=2, value="Product Title")
worksheet.cell(row=1, column=3, value="Amazon Price")

# Add the extracted data to the worksheet
for idx in range(2, len(data_asins) + 2):
    worksheet.cell(row=idx, column=1, value=data_asins[idx - 2])
    worksheet.cell(row=idx, column=2, value=product_titles[idx - 2])
    worksheet.cell(row=idx, column=3, value=prices[idx - 2])

workbook.save('amazon_products.xlsx')

workbook.close()

print("Data has been successfully saved to amazon_products.xlsx")